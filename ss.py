import random
import tkinter as tk
from tkinter import ttk
from turtle import st
from argon2 import Type
from openpyxl import Workbook
from openpyxl import load_workbook
from tomlkit import item

win = tk.Tk()

first_names = [
    "John",
    "Jane",
    "Michael",
    "Emily",
    "David",
    "Emma",
    "Christopher",
    "Olivia",
    "Daniel",
    "Sophia",
]
last_names = [
    "Smith",
    "Johnson",
    "Williams",
    "Jones",
    "Brown",
    "Davis",
    "Miller",
    "Wilson",
    "Moore",
    "Taylor",
]


def import_data():
    tree.delete(*tree.get_children())
    wb = load_workbook("test.xlsx")
    ws1 = wb.active
    data = []
    skip_first = True
    for row in ws1.iter_rows():
        if skip_first:
            skip_first = False
            continue
        for cell in row:
            data.append(cell.value)
        tree.insert(parent="", index="end", values=data)
        data.clear()


def export_data():
    wb = Workbook()
    wn = wb.active
    wn.cell(row=1, column=1, value="first name")
    wn.cell(row=1, column=2, value="last name")
    wn.cell(row=1, column=3, value="grade")
    r = 2
    for i in tree.get_children():
        col = 1
        for val in tree.item(i)["values"]:
            wn.cell(row=r, column=col, value=val)
            col = col + 1
        r = r + 1
    wb.save("./test.xlsx")


def create_data():
    n = len(last_names)
    for i in range(30):
        tree.insert(
            parent="",
            index="end",
            values=(
                first_names[int(random.random() * n)],
                last_names[int(random.random() * n)],
                int(random.random() * 100),
            ),
        )


win.geometry("1000x500")
tree = ttk.Treeview(
    win, columns=["fname", "lname", "grade"], selectmode="browse", show="headings"
)
scrol = tk.Scrollbar(win, command=tree.yview, orient="vertical")

tree.heading(column="fname", text="First Name")
tree.heading(column="lname", text="Last Name")
tree.heading(column="grade", text="Grade")
tree.bind("<Double-1>", lambda e: print(tree.focus()))
button = ttk.Button(win, text="generate data", command=create_data)
button1 = ttk.Button(win, text="export data", command=export_data)
button2 = ttk.Button(win, text="import data", command=import_data)

tree.configure(yscrollcommand=scrol.set)
scrol.grid(row=0, column=1, sticky="nswe")
tree.grid(row=0, column=0, sticky="se")
button.grid(row=1, column=0, sticky="se")
button1.grid(row=1, column=1, sticky="se")
button2.grid(row=1, column=2, sticky="se")

print(int(random.random() * 100))
win.mainloop()
